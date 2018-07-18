# -*- coding: utf-8 -*-
# 
import sys	
reload(sys)
sys.setdefaultencoding('utf8')
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import cell

# import assertpy
import unittest


i = 1

for first in dic.keys():
    j = 1

    if type(dic[first]) is not dict:
        now_sheet.cell(row=i + 1, column=j).value = first
        i += 1
    else:
        now_sheet.cell(row=i + 1, column=j).value = first
        i += 1
        for sec in dic[first].keys():
            j = 2

            if type(dic[first][sec]) is not dict:
                now_sheet.cell(row=i + 1, column=j).value = sec
                i += 1
            else:
                now_sheet.cell(row=i + 1, column=j).value = sec
                i += 1
                for thir in dic[first][sec].keys():
                    j = 3

                    if type(dic[first][sec][thir]) is not dict:
                        now_sheet.cell(row=i + 1, column=j).value = thir
                        i += 1
                    else:
                        now_sheet.cell(row=i + 1, column=j).value = thir
                        i += 1
                        for four in dic[first][sec][thir].keys():
                            j = 4

                            if type(dic[first][sec][thir][four]) is not dict:
                                now_sheet.cell(row=i + 1, column=j).value = four
                                i += 1
                            else:
                                now_sheet.cell(row=i + 1, column=j).value = four
                                i += 1
                                for fif in dic[first][sec][thir][four].keys():
                                    j = 5

                                    if type(dic[first][sec][thir][four][fif]) is not dict:
                                        now_sheet.cell(row=i + 1, column=j).value = fif
                                        i += 1
                                    else:
                                        now_sheet.cell(row=i + 1, column=j).value = fif
                                        i += 1
                                        for six in dic[first][sec][thir][four][fif].keys():
                                            j = 6

                                            if type(dic[first][sec][thir][four][fif][six]) is not dict:
                                                now_sheet.cell(row=i + 1, column=j).value = six
                                                i += 1
                                            else:
                                                now_sheet.cell(row=i + 1, column=j).value = six
                                                i += 1
                                                for seven in dic[first][sec][thir][four][fif][six].keys():
                                                    j = 7
                                                    now_sheet.cell(row=i + 1, column=j).value = seven
                                                    i += 1
a = wb.sheetnames
for i in a:
    if i == 'Sheet':
        del wb['Sheet']


wb.save(filename)	

# write_excel_title('Route1','/home/user/workspace/zh_nan/task/11.xlsx')


def compare_excel(inputfile,outputfile,l,mark):
    #比较源文件内容正确

    name = load_workbook(inputfile).sheetnames
    # print 'name',name
    type = True
    failsename = []
    for route in range(l*6-6,l*6):
        # print 'readin', route,l*6-6,l*6
        readin = load_workbook(inputfile)
        read_in = readin[name[route]]

        readout = load_workbook(outputfile)
        read_out = readout[name[route]]

        row = read_out.max_row
        column = read_out.max_column
        # print 'test',read_out.cell(row=2,column=1).value
        # print 'row,column',row,column
        for i in range(2,row+1):
            for j in range(1,column):
                if read_out.cell(row=i,column=j).value == read_in.cell(row=i,column=j).value:
                    read_out.cell(row=i,column=10).value = 'OK'
                    readout.save(outputfile)
                else:
                    read_out.cell(row=i,column=10).value = 'NG'
                    readout.save(outputfile)
                    type = False
                    break
        if type == False:
            failsename.append(name[route])
    print '%s Failsename :'%mark,failsename
    return type


# print compare_excel('/home/nan/zh_nan/task/outputexit.xlsx', '/home/nan/zh_nan/task/output.xlsx', 1)